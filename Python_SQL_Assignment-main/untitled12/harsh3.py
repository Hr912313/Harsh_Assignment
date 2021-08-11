import psycopg2
import xlrd
# from openpyxl.workbook import Workbook
import openpyxl
import os
import pandas as pd


class Departments:
    def create_tables(self):
        try:
            # read the connection parameters
            # connect to the PostgreSQL server
            conn = psycopg2.connect(
                host="localhost",
                database="Assignment1",
                user="postgres",
                password="912313")
            # Creating a cursor object using the cursor() method
            cur = conn.cursor()

            # reading xlsx file
            df = pd.read_excel("ques_2.xlsx")

            """ create tables in the PostgreSQL database"""
            # query for creating table
            commands = """
                CREATE TABLE Total_Compensation (
           ename varchar(10) ,
           empno numeric ,
           dname VARCHAR ( 50 ) ,
           total_compensation numeric,
           Months_Spent numeric )"""

            # query for required result

            query = """INSERT INTO Total_Compensation (ename, empno, dname, total_compensation, Months_Spent) VALUES (%s, %s, %s, %s, %s)"""


            # create table
            cur.execute(commands)

            for r in range(1, len(df)):
                # emp_name = sheet.cell(r,0).value
                ename = df['ename'][r]
                # emp_no= sheet.cell(r,1).value
                empno = int(df['empno'][r])
                # dept_name = sheet.cell(r,2).value
                dname = df['dname'][r]
                # Total_Compensation = sheet.cell(r,3).value
                total_compensation = int(df['total_compensation'][r])
                # Months_Spent = sheet.cell(r,4).value
                Months_Spent = int(df['months_spent'][r])

                values = (ename, empno, dname, total_compensation, Months_Spent)
                # inserting value one by one(row wise)
                cur.execute(query, values)

        except Exception as e:
            print("Error", e)
        finally:

            if conn is not None:
                cur.close()  # close communication with the PostgreSQL database server
                conn.commit()    # commit the changes



if __name__ == '__main__':
    department = Departments() # creating object
    department.create_tables() # calling method
